import calendar
import itertools
import os
import threading
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import wraps
from io import BytesIO

from dotenv import load_dotenv

load_dotenv()

import openpyxl
from flask import Flask, jsonify, render_template, request, send_file, session

import calculo
import calculo_defaults
import data_store
import mariadb_client
import parsers
from utils import normalizar, parse_data

app = Flask(__name__)
app.secret_key = os.environ["SECRET_KEY"]


def requer_login(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "papel" not in session:
            return jsonify({"error": "Não autenticado."}), 401
        return f(*args, **kwargs)
    return wrapper


def requer_papel(*papeis_permitidos):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            papel = session.get("papel")
            if not papel:
                return jsonify({"error": "Não autenticado."}), 401
            if papel not in papeis_permitidos:
                return jsonify({"error": "Sem permissão para esta ação."}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/status")
def status():
    return jsonify({"ok": True})


def _usuario_sessao_json():
    return {"username": session["username"], "papel": session["papel"], "nome": session.get("nome") or ""}


@app.route("/api/precisa-bootstrap")
def precisa_bootstrap():
    return jsonify({"precisaBootstrap": not data_store.existe_algum_usuario()})


@app.route("/api/bootstrap", methods=["POST"])
def bootstrap():
    if data_store.existe_algum_usuario():
        return jsonify({"error": "Sistema já tem usuários cadastrados."}), 403
    payload = request.get_json(force=True)
    username = (payload.get("username") or "").strip()
    senha = payload.get("senha") or ""
    nome = (payload.get("nome") or "").strip()
    if not username or len(senha) < 6:
        return jsonify({"error": "Informe usuário e senha (mínimo 6 caracteres)."}), 400
    usuario = data_store.criar_usuario(username, senha, "diretoria", nome)
    session["user_id"] = usuario["id"]
    session["username"] = usuario["username"]
    session["papel"] = usuario["papel"]
    session["nome"] = usuario["nome"]
    return jsonify(_usuario_sessao_json())


@app.route("/api/login", methods=["POST"])
def login():
    payload = request.get_json(force=True)
    username = (payload.get("username") or "").strip()
    senha = payload.get("senha") or ""
    usuario = data_store.autenticar(username, senha)
    if not usuario:
        return jsonify({"error": "Usuário ou senha inválidos."}), 401
    session["user_id"] = usuario["id"]
    session["username"] = usuario["username"]
    session["papel"] = usuario["papel"]
    session["nome"] = usuario["nome"]
    return jsonify(_usuario_sessao_json())


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return "", 204


@app.route("/api/me")
def me():
    if "papel" not in session:
        return jsonify({"error": "Não autenticado."}), 401
    return jsonify(_usuario_sessao_json())


def _ler_matriz(arquivo):
    """Le um arquivo .xlsx enviado (multipart) e devolve a primeira aba como
    lista de listas, igual ao sheet_to_json({header:1}) do SheetJS."""
    wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    matriz = [[("" if v is None else v) for v in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return matriz


def _funcionario_json(f):
    d = dict(f)
    if isinstance(d.get("admissao"), date):
        d["admissao"] = d["admissao"].isoformat()
    return d


def _pesagem_json(p):
    d = dict(p)
    if isinstance(d.get("data"), date):
        d["data"] = d["data"].isoformat()
    return d


@app.route("/api/funcionarios/atualizar", methods=["POST"])
@requer_papel("coordenador", "gerente", "diretoria")
def atualizar_funcionarios():
    try:
        colunas, linhas = mariadb_client.buscar_funcionarios()
        lista = parsers.parse_funcionarios_mariadb(colunas, linhas, data_store.get_dias_base())
    except Exception as exc:
        return jsonify({"error": f"Falha ao buscar funcionários no sistema da empresa: {exc}"}), 400
    data_store.set_funcionarios(lista)
    return jsonify({"ok": True, "total": len(lista)})


@app.route("/api/import/frotas", methods=["POST"])
@requer_papel("coordenador", "gerente", "diretoria")
def import_frotas():
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400
    try:
        matriz = _ler_matriz(arquivo)
        lista = parsers.parse_frotas(matriz)
    except parsers.ErroImportacao as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": f"Falha ao importar disponibilidade: {exc}"}), 400
    data_store.set_frotas(lista)
    return jsonify({"ok": True, "total": len(lista), "nomeArquivo": arquivo.filename})


@app.route("/api/import/pesagens", methods=["POST"])
@requer_papel("coordenador", "gerente", "diretoria")
def import_pesagens():
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return jsonify({"error": "Nenhum arquivo enviado."}), 400
    try:
        matriz = _ler_matriz(arquivo)
        resultado = parsers.parse_pesagens(matriz)
    except parsers.ErroImportacao as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": f"Falha ao importar produções: {exc}"}), 400
    data_store.set_pesagens(resultado["regs"])
    return jsonify({
        "ok": True,
        "total": len(resultado["regs"]),
        "temKm": resultado["tem_km"],
        "colunas": resultado["colunas"],
        "nomeArquivo": arquivo.filename,
    })


def _processar_jornada_async(periodo_inicio, periodo_fim):
    """Roda em thread separada - buscar e processar a jornada (100k+ linhas)
    dentro da própria requisição HTTP estoura o timeout fixo do proxy do
    Render (~30s). Por isso a rota devolve na hora e isso aqui atualiza o
    status que o front fica consultando."""
    try:
        colunas, linhas = mariadb_client.buscar_jornada(periodo_inicio, periodo_fim)
        mapa = parsers.parse_jornada(itertools.chain([colunas], linhas))
        # guarda so a CONTAGEM de dias por matricula (nao a lista inteira de
        # datas - isso e' 100k+ strings sem necessidade) junto com o periodo
        # em que foi buscado; calculo.py so usa esse dado se o periodo bater
        # com o periodo de apuracao atual, senao ignora (mais seguro do que
        # filtrar datas de um periodo antigo contra um periodo novo).
        compacto = {
            "periodoInicio": periodo_inicio,
            "periodoFim": periodo_fim,
            "porMat": {mat: {"totalDias": len(v["dias"]), "faltas": v["faltas"]} for mat, v in mapa.items()},
        }
        data_store.set_jornada(compacto)
        total_registros = sum(len(v["faltas"]) for v in compacto["porMat"].values())
        data_store.set_jornada_import_status({
            "status": "concluido", "totalMatriculas": len(mapa), "totalRegistros": total_registros,
            "periodoInicio": periodo_inicio, "periodoFim": periodo_fim,
        })
    except parsers.ErroImportacao as exc:
        data_store.set_jornada_import_status({"status": "erro", "mensagem": str(exc)})
    except Exception as exc:
        data_store.set_jornada_import_status({"status": "erro", "mensagem": f"Falha ao buscar jornada no sistema da empresa: {exc}"})


@app.route("/api/jornada/atualizar", methods=["POST"])
@requer_papel("coordenador", "gerente", "diretoria")
def atualizar_jornada():
    periodo = data_store.get_periodo()
    inicio, fim = periodo.get("inicio"), periodo.get("fim")
    if not inicio or not fim:
        return jsonify({"error": "Período de apuração inválido."}), 400
    data_store.set_jornada_import_status({"status": "processando"})
    threading.Thread(target=_processar_jornada_async, args=(inicio, fim), daemon=True).start()
    return jsonify({"status": "processando"}), 202


@app.route("/api/import/jornada/status", methods=["GET"])
@requer_login
def status_import_jornada():
    return jsonify(data_store.get_jornada_import_status())


@app.route("/api/seed", methods=["POST"])
@requer_papel("coordenador", "gerente", "diretoria")
def seed_demo():
    periodo = data_store.get_periodo()
    inicio = parse_data(periodo["inicio"])
    fim = parse_data(periodo["fim"])
    if not inicio or not fim:
        return jsonify({"error": "Período inválido."}), 400
    gerado = parsers.gerar_dados_demo(inicio, fim)
    data_store.set_funcionarios(gerado["funcionarios"])
    data_store.set_pesagens(gerado["pesagens"])
    return jsonify({"ok": True, "totalFuncionarios": len(gerado["funcionarios"]), "totalPesagens": len(gerado["pesagens"])})


@app.route("/api/dados", methods=["GET"])
@requer_login
def get_dados():
    d = data_store.get_dados_bundle()
    return jsonify({
        "funcionarios": [_funcionario_json(f) for f in d["funcionarios"]],
        "pesagens": [_pesagem_json(p) for p in d["pesagens"]],
        "frotas": d["frotas"],
        "periodo": d["periodo"],
        "diasBase": d["dias_base"],
        "jornadaResumo": d["jornada_resumo"],
        "funcionariosAtualizadoEm": d["funcionarios_atualizado_em"],
    })


@app.route("/api/dados", methods=["DELETE"])
@requer_papel("gerente", "diretoria")
def limpar_dados():
    data_store.limpar_tudo()
    return "", 204


@app.route("/api/dados/<base>", methods=["DELETE"])
@requer_papel("gerente", "diretoria")
def limpar_uma_base(base):
    """Zera so uma base (ex.: pesagens) pra permitir subir um arquivo novo sem
    derrubar as outras."""
    if base not in data_store.BASES_LIMPAVEIS:
        return jsonify({"error": "Base desconhecida."}), 400
    data_store.limpar_base(base)
    return "", 204


@app.route("/api/periodo", methods=["PUT"])
@requer_papel("gerente", "diretoria")
def set_periodo():
    payload = request.get_json(force=True)
    inicio = payload.get("inicio")
    fim = payload.get("fim")
    dias_base = payload.get("diasBase")
    if not inicio or not fim:
        return jsonify({"error": "Informe início e fim do período."}), 400
    if not parse_data(inicio) or not parse_data(fim):
        return jsonify({"error": "Datas inválidas."}), 400
    data_store.set_periodo(inicio, fim)
    if dias_base:
        data_store.set_dias_base(int(dias_base))
    return jsonify({"periodo": data_store.get_periodo(), "diasBase": data_store.get_dias_base()})


@app.route("/api/parametros", methods=["GET"])
@requer_login
def get_parametros():
    return jsonify(data_store.get_parametros())


@app.route("/api/parametros", methods=["PUT"])
@requer_papel("gerente", "diretoria")
def set_parametros():
    payload = request.get_json(force=True)
    data_store.set_parametros(payload)
    return jsonify(data_store.get_parametros())


@app.route("/api/parametros/restaurar-faixas", methods=["POST"])
@requer_papel("gerente", "diretoria")
def restaurar_faixas():
    atual = data_store.get_parametros()
    atual.update(calculo_defaults.faixas_padrao())
    data_store.set_parametros(atual)
    return jsonify(atual)


def _semana_json(sem):
    d = dict(sem)
    d.pop("_vCod", None)
    d["dias"] = {iso: {k: v for k, v in dia.items() if k != "_vCod"} for iso, dia in d.get("dias", {}).items()}
    return d


def _agregado_json(k):
    d = dict(k)
    d.pop("_vCod", None)
    if isinstance(d.get("admissao"), date):
        d["admissao"] = d["admissao"].isoformat()
    d["semanas"] = {str(idx): _semana_json(sem) for idx, sem in d.get("semanas", {}).items()}
    return d


def _calcular_atual():
    """Roda calculo.calcular() com os dados/parametros/ajustes atuais. Devolve
    None (nos dois campos) se o periodo estiver invalido."""
    d = data_store.get_calculo_inputs()
    inicio = parse_data(d["periodo"]["inicio"])
    fim = parse_data(d["periodo"]["fim"])
    if not inicio or not fim:
        return None, d
    resultado = calculo.calcular(
        d["funcionarios"], d["pesagens"], inicio, fim, d["dias_base"], d["parametros"], d["ajustes"], d["jornada"],
    )
    return resultado, d


@app.route("/api/calculo", methods=["GET"])
@requer_login
def get_calculo():
    resultado, _ = _calcular_atual()
    if resultado is None:
        return jsonify({"error": "Período inválido."}), 400
    return jsonify({
        "lista": [_agregado_json(k) for k in resultado["lista"]],
        "nSem": resultado["nSem"],
    })


_ESPEC_LABEL = {"CAMINHAO": "Caminhão", "BATE-VOLTA": "Bate-volta", "COLHEDORA": "Colhedora", "TRANSBORDO": "Transbordo"}


@app.route("/api/calculo/exportar", methods=["GET"])
@requer_login
def exportar_calculo_xlsx():
    resultado, d = _calcular_atual()
    if resultado is None:
        return jsonify({"error": "Período inválido."}), 400

    # a tela permite fixar varios colaboradores no filtro, entao vem um "busca"
    # por termo; a linha entra se casar com qualquer um deles
    buscas = [normalizar(b) for b in request.args.getlist("busca") if b.strip()]
    apenas_com_producao = request.args.get("apenasComProducao", "1") != "0"
    especialidade = request.args.get("especialidade", "TODOS")
    departamento = request.args.get("departamento", "TODOS")
    exibir_nomes = request.args.get("nomes", "0") == "1"

    lista = [
        k for k in resultado["lista"]
        if (not apenas_com_producao or k["viagens"] > 0)
        and (especialidade == "TODOS" or k["espec"] == especialidade)
        and (departamento == "TODOS" or (k.get("departamento") or "Não informado") == departamento)
        and (not buscas or any(b in normalizar(k.get("nome")) or b in str(k["mat"]) for b in buscas))
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cálculo"
    cabecalho = ["Matrícula"] + (["Nome"] if exibir_nomes else []) + [
        "Departamento", "Admissão", "Especialidade", "Dias", "Dias trabalhados", "Dias sem expediente",
        "Viagens", "Toneladas", "Km méd.",
        "Salário base (R$)", "Gratificação (R$)", "Teto (R$)", "% Atingido", "Ajuste manual (%)", "Total a receber (R$)",
    ]
    ws.append(cabecalho)
    for k in lista:
        admissao = k.get("admissao")
        linha = [k["mat"]] + ([k.get("nome", "")] if exibir_nomes else []) + [
            k.get("departamento") or "Não informado",
            admissao.isoformat() if isinstance(admissao, date) else "",
            _ESPEC_LABEL.get(k["espec"], k["espec"]),
            k.get("diasTrabalhados", k.get("dias")),
            k.get("diasTrabalhadosReal", k.get("diasTrabalhados", k.get("dias"))),
            len(k.get("faltas") or []),
            k["viagens"],
            round(k["ton"], 1),
            round(k.get("kmMed") or 0, 0),
            round(k["sal"], 2),
            round(k["gratif"], 2),
            round(k.get("tetoEfetivo", k["teto"]), 0),
            round(k["atingPct"] * 100, 1),
            round(k["ajustePct"], 1) if k.get("ajustePct") else "",
            round(k["totalReceber"], 2),
        ]
        ws.append(linha)

    for col_cells in ws.columns:
        largura = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(largura, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome_arquivo = f"calculo_gratificacao_{d['periodo']['inicio']}_a_{d['periodo']['fim']}.xlsx"
    return send_file(
        buf, as_attachment=True, download_name=nome_arquivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/ajuste/<mat>", methods=["PUT"])
@requer_papel("gerente", "diretoria")
def set_ajuste(mat):
    payload = request.get_json(force=True)
    pct = payload.get("pct")
    obs = (payload.get("obs") or "").strip()
    if pct is None:
        return jsonify({"error": "Informe o percentual do ajuste."}), 400
    if not obs:
        return jsonify({"error": "Informe a observação justificando o ajuste."}), 400
    data_store.set_ajuste(mat, float(pct), obs)
    return jsonify({"ok": True})


@app.route("/api/ajuste/<mat>", methods=["DELETE"])
@requer_papel("gerente", "diretoria")
def remover_ajuste(mat):
    data_store.remover_ajuste(mat)
    return "", 204


def _ajuste_pendente_json(row):
    d = dict(row)
    for campo in ("criado_em", "aprovado_gerente_em", "aprovado_diretoria_em", "rejeitado_em"):
        if isinstance(d.get(campo), datetime):
            d[campo] = d[campo].isoformat()
    if isinstance(d.get("pct"), Decimal):
        d["pct"] = float(d["pct"])
    return {
        "id": d["id"],
        "mat": d["mat"],
        "pct": d["pct"],
        "obs": d["obs"],
        "criadoPor": d["criado_por"],
        "criadoEm": d["criado_em"],
        "status": d["status"],
        "aprovadoGerente": d["aprovado_gerente"],
        "aprovadoGerentePor": d["aprovado_gerente_por"],
        "aprovadoDiretoria": d["aprovado_diretoria"],
        "aprovadoDiretoriaPor": d["aprovado_diretoria_por"],
        "rejeitadoPor": d["rejeitado_por"],
        "motivoRejeicao": d["motivo_rejeicao"],
    }


@app.route("/api/ajustes-pendentes", methods=["GET"])
@requer_papel("gerente", "diretoria")
def listar_ajustes_pendentes():
    return jsonify([_ajuste_pendente_json(r) for r in data_store.listar_ajustes_pendentes()])


@app.route("/api/ajustes-pendentes", methods=["POST"])
@requer_papel("coordenador")
def criar_ajuste_pendente():
    payload = request.get_json(force=True)
    mat = (payload.get("mat") or "").strip()
    pct = payload.get("pct")
    obs = (payload.get("obs") or "").strip()
    if not mat or pct is None:
        return jsonify({"error": "Informe o colaborador e o percentual."}), 400
    if not obs:
        return jsonify({"error": "Informe a observação justificando o ajuste."}), 400
    row = data_store.criar_ajuste_pendente(mat, float(pct), obs, session["username"])
    return jsonify(_ajuste_pendente_json(row))


@app.route("/api/ajustes-pendentes/aprovar", methods=["POST"])
@requer_papel("gerente", "diretoria")
def aprovar_ajustes_pendentes():
    payload = request.get_json(force=True)
    ids = payload.get("ids") or []
    if not ids:
        return jsonify({"error": "Selecione ao menos um ajuste."}), 400
    aplicados = data_store.aprovar_ajustes(ids, session["papel"], session["username"])
    return jsonify({"ok": True, "aplicados": aplicados})


@app.route("/api/ajustes-pendentes/<int:ajuste_id>/rejeitar", methods=["POST"])
@requer_papel("gerente", "diretoria")
def rejeitar_ajuste_pendente(ajuste_id):
    payload = request.get_json(force=True)
    motivo = (payload.get("motivo") or "").strip()
    if not motivo:
        return jsonify({"error": "Informe o motivo da rejeição."}), 400
    data_store.rejeitar_ajuste(ajuste_id, session["username"], motivo)
    return "", 204


@app.route("/api/gratificacoes/aprovacoes", methods=["GET"])
@requer_papel("gerente", "diretoria")
def listar_aprovacoes_gratificacao():
    periodo = data_store.get_periodo()
    aprovacoes = data_store.listar_aprovacoes_gratificacao(periodo["inicio"], periodo["fim"])
    resultado = {}
    for mat, row in aprovacoes.items():
        resultado[mat] = {
            "aprovadoGerente": row["aprovado_gerente"],
            "aprovadoGerentePor": row["aprovado_gerente_por"],
            "aprovadoDiretoria": row["aprovado_diretoria"],
            "aprovadoDiretoriaPor": row["aprovado_diretoria_por"],
        }
    return jsonify(resultado)


def _proximo_periodo(fim):
    """O periodo seguinte comeca no dia seguinte ao fim e termina no mesmo dia
    do mes seguinte: 16/06-15/07 fecha e abre 16/07-15/08."""
    ano, mes = (fim.year + 1, 1) if fim.month == 12 else (fim.year, fim.month + 1)
    ultimo_dia = calendar.monthrange(ano, mes)[1]
    return fim + timedelta(days=1), date(ano, mes, min(fim.day, ultimo_dia))


@app.route("/api/fechamentos", methods=["GET"])
@requer_login
def listar_fechamentos():
    return jsonify(data_store.get_fechamentos_indice())


@app.route("/api/fechamentos/<inicio>/<fim>", methods=["GET"])
@requer_login
def detalhe_fechamento(inicio, fim):
    registro = data_store.get_fechamento(inicio, fim)
    if not registro:
        return jsonify({"error": "Período fechado não encontrado."}), 404
    return jsonify(registro)


@app.route("/api/fechamentos/<inicio>/<fim>/enriquecer", methods=["POST"])
@requer_papel("gerente", "diretoria")
def enriquecer_fechamento(inicio, fim):
    """Completa um periodo ja fechado com o detalhe que o snapshot antigo nao
    guardava (semanas, frotas, faltas), recalculando a partir das pesagens
    reimportadas.

    NUNCA mexe no dinheiro: gratificacao, teto, % atingido, ajuste manual e
    total a receber continuam sendo os valores congelados no fechamento. Isso
    aqui so acrescenta a quebra que faltava pro extrato do historico."""
    registro = data_store.get_fechamento(inicio, fim)
    if not registro:
        return jsonify({"error": "Período fechado não encontrado."}), 404

    periodo_atual = data_store.get_periodo()
    if periodo_atual["inicio"] != inicio or periodo_atual["fim"] != fim:
        return jsonify({
            "error": f"Ajuste o período de apuração para {inicio} a {fim} antes de completar o detalhe "
                     "(o cálculo usa o período atual para montar as semanas)."
        }), 400

    resultado, d = _calcular_atual()
    if resultado is None:
        return jsonify({"error": "Período inválido."}), 400
    if not d["pesagens"]:
        return jsonify({"error": "Nenhuma pesagem carregada — reimporte a planilha do período antes."}), 400

    por_mat = {str(k["mat"]): k for k in resultado["lista"]}
    completados, sem_dados = 0, 0
    for item in registro["colaboradores"]:
        k = por_mat.get(str(item["mat"]))
        if not k:
            sem_dados += 1
            continue
        item["semanas"] = {str(idx): _semana_json(sem) for idx, sem in (k.get("semanas") or {}).items()}
        item["frotas"] = k.get("frotas") or {}
        item["faltas"] = k.get("faltas") or []
        if item.get("kmMed") in (None, 0):
            item["kmMed"] = k.get("kmMed")
        if item.get("prodR$") in (None, 0):
            item["prodR$"] = k.get("prodR$")
        if not item.get("admissao") and isinstance(k.get("admissao"), date):
            item["admissao"] = k["admissao"].isoformat()
        completados += 1

    registro["nSem"] = resultado["nSem"]
    registro["detalheCompletadoEm"] = datetime.now().isoformat(timespec="seconds")
    registro["detalheCompletadoPor"] = session["username"]
    data_store.salvar_fechamento(registro)
    return jsonify({"ok": True, "completados": completados, "semDados": sem_dados})


@app.route("/api/fechamento", methods=["POST"])
@requer_papel("diretoria")
def fechar_periodo():
    """Congela a gratificacao do periodo atual num snapshot, limpa os dados que
    eram daquele periodo (pesagens, frotas, jornada e ajustes manuais) e avanca
    o periodo de apuracao - deixando a tela pronta pra importar o mes seguinte.
    As aprovacoes ficam onde estao: a tabela ja e escopada por periodo."""
    resultado, d = _calcular_atual()
    if resultado is None:
        return jsonify({"error": "Período inválido."}), 400
    periodo = d["periodo"]
    fim = parse_data(periodo["fim"])

    linhas = [
        k for k in resultado["lista"]
        if k["gratif"] > 0 and (k["viagens"] > 0 or k.get("ajustePct"))
    ]
    if not linhas:
        return jsonify({"error": "Não há gratificação calculada neste período para fechar."}), 400

    aprovacoes = data_store.listar_aprovacoes_gratificacao(periodo["inicio"], periodo["fim"])
    # guarda tambem semanas/frotas/faltas pra o extrato do periodo fechado ficar
    # tao completo quanto o do periodo corrente
    campos = ("mat", "nome", "funcao", "departamento", "espec", "viagens", "ton", "kmMed", "sal",
              "gratif", "teto", "tetoEfetivo", "atingPct", "ajustePct", "ajusteObs", "ajusteValor",
              "totalReceber", "diasPeriodo", "diasTrabalhados", "diasTrabalhadosReal",
              "prodR$", "frotas", "faltas", "admissao")
    colaboradores = []
    for k in linhas:
        item = {c: k.get(c) for c in campos}
        if isinstance(item.get("admissao"), date):
            item["admissao"] = item["admissao"].isoformat()
        item["semanas"] = {str(idx): _semana_json(sem) for idx, sem in (k.get("semanas") or {}).items()}
        ap = aprovacoes.get(str(k["mat"])) or {}
        item["aprovadoGerentePor"] = ap.get("aprovado_gerente_por")
        item["aprovadoDiretoriaPor"] = ap.get("aprovado_diretoria_por")
        colaboradores.append(item)

    somar = lambda campo: round(sum(k.get(campo) or 0 for k in linhas), 2)
    resumo = {
        "colaboradores": len(colaboradores),
        "viagens": sum(k["viagens"] for k in linhas),
        "ton": somar("ton"),
        "gratif": somar("gratif"),
        "teto": somar("tetoEfetivo"),
        "ajusteValor": somar("ajusteValor"),
        "sal": somar("sal"),
        "totalReceber": somar("totalReceber"),
    }
    registro = {
        "periodo": periodo,
        "diasBase": d["dias_base"],
        "nSem": resultado["nSem"],
        "fechadoEm": datetime.now().isoformat(timespec="seconds"),
        "fechadoPor": session["username"],
        "parametros": d["parametros"],
        "resumo": resumo,
        "colaboradores": colaboradores,
    }
    data_store.salvar_fechamento(registro)

    for base in ("pesagens", "frotas", "jornada"):
        data_store.limpar_base(base)
    data_store.limpar_ajustes()
    data_store.descartar_ajustes_pendentes()

    novo_inicio, novo_fim = _proximo_periodo(fim)
    data_store.set_periodo(novo_inicio.isoformat(), novo_fim.isoformat())
    return jsonify({
        "ok": True,
        "resumo": resumo,
        "periodoFechado": periodo,
        "novoPeriodo": {"inicio": novo_inicio.isoformat(), "fim": novo_fim.isoformat()},
    })


@app.route("/api/gratificacoes/aprovar", methods=["POST"])
@requer_papel("gerente", "diretoria")
def aprovar_gratificacoes_rota():
    payload = request.get_json(force=True)
    mats = payload.get("mats") or []
    if not mats:
        return jsonify({"error": "Selecione ao menos um colaborador."}), 400
    periodo = data_store.get_periodo()
    total = data_store.aprovar_gratificacoes(mats, session["papel"], session["username"], periodo["inicio"], periodo["fim"])
    return jsonify({"ok": True, "total": total})


@app.route("/api/gratificacoes/desfazer", methods=["POST"])
@requer_papel("gerente", "diretoria")
def desfazer_aprovacao_gratificacao_rota():
    payload = request.get_json(force=True)
    mats = payload.get("mats") or []
    if not mats:
        return jsonify({"error": "Selecione ao menos um colaborador."}), 400
    periodo = data_store.get_periodo()
    total = data_store.desfazer_aprovacao_gratificacao(mats, session["papel"], periodo["inicio"], periodo["fim"])
    return jsonify({"ok": True, "total": total})


def _usuario_json(u):
    d = dict(u)
    if isinstance(d.get("criado_em"), datetime):
        d["criado_em"] = d["criado_em"].isoformat()
    return d


@app.route("/api/usuarios", methods=["GET"])
@requer_papel("diretoria")
def listar_usuarios():
    return jsonify([_usuario_json(u) for u in data_store.listar_usuarios()])


@app.route("/api/usuarios", methods=["POST"])
@requer_papel("diretoria")
def criar_usuario_rota():
    payload = request.get_json(force=True)
    username = (payload.get("username") or "").strip()
    senha = payload.get("senha") or ""
    papel = payload.get("papel")
    nome = (payload.get("nome") or "").strip()
    if not username or len(senha) < 6:
        return jsonify({"error": "Informe usuário e senha (mínimo 6 caracteres)."}), 400
    if papel not in data_store.PAPEIS:
        return jsonify({"error": "Papel inválido."}), 400
    if data_store.buscar_usuario_por_username(username):
        return jsonify({"error": "Já existe um usuário com esse nome."}), 400
    return jsonify(_usuario_json(data_store.criar_usuario(username, senha, papel, nome)))


@app.route("/api/usuarios/<int:user_id>", methods=["PUT"])
@requer_papel("diretoria")
def atualizar_usuario_rota(user_id):
    payload = request.get_json(force=True)
    papel = payload.get("papel")
    if papel is not None and papel not in data_store.PAPEIS:
        return jsonify({"error": "Papel inválido."}), 400
    senha = payload.get("senha") or None
    if senha and len(senha) < 6:
        return jsonify({"error": "Senha deve ter ao menos 6 caracteres."}), 400
    data_store.atualizar_usuario(
        user_id, papel=papel, nome=payload.get("nome"), ativo=payload.get("ativo"), senha=senha,
    )
    return jsonify({"ok": True})


@app.route("/api/usuarios/<int:user_id>", methods=["DELETE"])
@requer_papel("diretoria")
def remover_usuario_rota(user_id):
    if user_id == session.get("user_id"):
        return jsonify({"error": "Você não pode remover seu próprio usuário."}), 400
    data_store.remover_usuario(user_id)
    return "", 204


if __name__ == "__main__":
    app.run(debug=True, port=5002)
